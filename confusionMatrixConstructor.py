import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from imageSimilarity import image_similarity
from utils import text_to_images

characters = [
    "A", "Ą", "B", "C", "Ć", "D", "E", "Ę", "F", "G", "H", "I", "J", "K", "L", "Ł", "M", "N", "Ń", "O", "Ó", "P", "Q",
    "R", "S", "Ś", "T", "U", "V", "W", "X", "Y", "Z", "Ż", "Ź", "a", "ą", "b", "c", "ć", "d", "e", "ę", "f", "g", "h",
    "i", "j", "k", "l", "ł", "m", "n", "ń", "o", "ó", "p", "q", "r", "s", "ś", "t", "u", "v", "w", "x", "y", "z", "ż",
    "ź", "1", "2", "3", "4", "5", "6", "7", "8", "9", "0", "rn", "cl", "cI", "rt", "vv", "ri", "13"
]


def construct_confusion_matrix():
    headers = [""] + characters
    char_count = len(characters)
    data_2d = [[0] * char_count for _ in range(char_count)]

    for i, char in enumerate(characters):
        data_2d[0][i] = char
        data = 
        for j in range(i + 1):
            if i == j:
                data_2d[i][j] = 1
                continue

            images = text_to_images(characters[i], characters[j])
            img1 = images[0]
            img2 = images[1]

            similarity_score = image_similarity(img1, img2)
            data_2d[i][j] = similarity_score
            data_2d[j][i] = similarity_score

# Convert list to DataFrame
    df = pd.DataFrame(data_2d)

    # Save to Excel
    excel_path = "confusion_matrix.xlsx"
    df.to_excel(excel_path, index=False, header=False, engine="openpyxl")

    # Load workbook and select active sheet
    wb = load_workbook(excel_path)
    ws = wb.active

    # Get last column and row for table reference
    last_col = get_column_letter(char_count + 1)
    last_row = char_count + 1

    # Create formatted table
    table = Table(displayName="ConfusionMatrix", ref=f"A1:{last_col}{last_row}")

    # Apply table style
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=True,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save formatted Excel file
    wb.save(excel_path)

    print(f"Confusion matrix saved to '{excel_path}'!")


